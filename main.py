from vk_api import VkApi
from vk_api.bot_longpoll import VkBotLongPoll, VkBotEventType
from vk_api.utils import get_random_id
from openpyxl import Workbook, load_workbook
from datetime import datetime


VERSION = "5.1.1"

"""-------------------------------------------Глобальные константы-------------------------------------------"""


TOKEN = "vk1.a.fZnFDaMgklg5QhCrUZ2kEwkEksR32ZQ4FWg-5mEvzZbFN2Q-k1D-HA6ioVOTvUw7Pyz_" \
        "B25cw9ezloK3tmobWizeuhuaYsaxnZ8JL3Gs5k-E5ApXFtAcsxw-BzoJo6710kYE0Tte4GxQVaFZTbhjr-" \
        "KmQeh_P9JQeIeEKmtdJbPA4_MZxxHtXAkeW2IjNBCjSj_MZe7s5038sW2f_3lpmg"            # Токен группы для API
GROUP_ID = 186471220                                    # ID группы
ADMIN_ID = [286001249, 152553720]                       # ID администраторов-юзеров в беседе
CHAT_ID = 2                                             # ID беседы с ботом (Закупка снаряги + спам)
CHAT_ID_KICK = 3                                        # ID беседы для кика (Беседа новобранцев)
PERCENT = 20                                            # Требования к посещению (процент)
NAME_TABLE = 'G:/Мой диск/Радиус Журнал.xlsx'           # Название файла с таблицей
DATES_ROW = 1                                           # Номер строки с датами в таблице
USERS_COLUMN = 2                                        # Номер столбца с именами пользователей в таблице
LAST_DATE_COLUMN = 12                                   # Номер столбца с последней датой в таблице

authorize = VkApi(token=TOKEN)                          # Авторизация бота
session_api = authorize.get_api()                       # Авторизация сессии
longpool = VkBotLongPoll(authorize, group_id=GROUP_ID)  # Вход под LongPoll от лица группы


"""-----------------------------------------------Блок функций-----------------------------------------------"""


def open_table(data_only: bool) -> Workbook:
    """
    Функция открывает существующую таблицу для записи
    :param data_only: флаг преобразования данных по формуле или чистые данные **(controls whether cells with formulae
                        have either the formula (default) or the value stored the last time Excel read the sheet)**
    :return: объект xlsx таблицы
    """
    try:
        workbook = load_workbook(NAME_TABLE, data_only=data_only)
    except FileNotFoundError:
        print('Файл таблицы не найден!')
        workbook = Workbook()
    return workbook


def close_table(workbook: Workbook) -> None:
    """
    Функция закрывает объект таблицы с сохранением данных
    :param workbook: объект xlsx таблицы
    :return:
    """
    workbook.save(NAME_TABLE)


def update_table_date(workbook: Workbook, date: str) -> None:
    """
    Функция обновляет даты в таблице: создание новой колонки с датой при ее отсутствии в таблице для последующей записи
    :param workbook: объект xlsx таблицы
    :param date: наименование даты в формате строки
    :return:
    """
    sheet = workbook['Посещаемость']                                        # Лист таблицы "Посещаемость"
    date_list = list(sheet.iter_rows(min_row=DATES_ROW, max_row=DATES_ROW, values_only=True))[0]      # Список всех дат

    if date not in date_list:                                               # Проверка новой даты
        sheet.insert_cols(LAST_DATE_COLUMN)                                 # Создаем новую дату при необходимости
        sheet.cell(row=DATES_ROW, column=LAST_DATE_COLUMN).value = date


def update_table_user(user_id: str, user: str, fio: str) -> None:
    """
    Функция обновляет данные о пользователе в таблице. Производится установка или обновление данных о пользователе:
    ID, наименование, ФИО, отметка о посещении.
    :param user_id: ID юзера в формате строки
    :param user: наименование юзера (по ВК) в формате строки
    :param fio: ФИО юзера в формате строки
    :return:
    """
    workbook = open_table(data_only=False)

    sheet_1 = workbook['Посещаемость']                                                      # Лист 1
    sheet_2 = workbook['Соревнования и тренировки']                                         # Лист 2
    sheet_3 = workbook['Вооружение и транспорт']                                            # Лист 3
    sheet_4 = workbook['Данные об экипировке']                                              # Лист 4

    user_id_list = list(sheet_1.iter_cols(min_col=USERS_COLUMN - 1, max_col=USERS_COLUMN - 1, values_only=True))[0]

    if user_id in user_id_list:                                                             # Проверка нового юзера
        user_index_1 = user_id_list.index(user_id) + 1                                      # Установка индекса 1 листа
    else:
        sheet_1.cell(row=sheet_1.max_row + 1, column=USERS_COLUMN - 1).value = user_id      # Запись ID юзера в 1 лист
        sheet_1.cell(row=sheet_1.max_row, column=USERS_COLUMN).value = user                 # Запись юзера в 1 лист
        sheet_1.cell(row=sheet_1.max_row, column=USERS_COLUMN + 1).value = fio              # Запись ФИО в 1 лист
        sheet_1.cell(row=sheet_1.max_row, column=USERS_COLUMN + 2).value = '7.Рекрут'       # Запись звания в 1 лист

        sheet_2.cell(row=sheet_1.max_row + 1, column=USERS_COLUMN - 1).value = user_id      # Запись ID юзера в 2 лист
        sheet_2.cell(row=sheet_1.max_row, column=USERS_COLUMN).value = user                 # Запись юзера в 2 лист
        sheet_2.cell(row=sheet_1.max_row, column=USERS_COLUMN + 1).value = fio              # Запись ФИО в 2 лист
        sheet_2.cell(row=sheet_1.max_row, column=USERS_COLUMN + 2).value = '7.Рекрут'       # Запись звания в 2 лист

        sheet_3.cell(row=sheet_1.max_row + 1, column=USERS_COLUMN - 1).value = user_id      # Запись ID юзера в 3 лист
        sheet_3.cell(row=sheet_1.max_row, column=USERS_COLUMN).value = user                 # Запись юзера в 3 лист
        sheet_3.cell(row=sheet_1.max_row, column=USERS_COLUMN + 1).value = fio              # Запись ФИО в 3 лист
        sheet_3.cell(row=sheet_1.max_row, column=USERS_COLUMN + 2).value = '7.Рекрут'       # Запись звания в 3 лист

        sheet_4.cell(row=sheet_2.max_row + 1, column=USERS_COLUMN - 1).value = user_id      # Запись ID юзера в 4 лист
        sheet_4.cell(row=sheet_2.max_row, column=USERS_COLUMN).value = user                 # Запись юзера в 4 лист
        sheet_4.cell(row=sheet_1.max_row, column=USERS_COLUMN + 1).value = fio              # Запись ФИО в 4 лист
        sheet_4.cell(row=sheet_1.max_row, column=USERS_COLUMN + 2).value = '7.Рекрут'       # Запись звания в 4 лист

        user_index_1 = sheet_1.max_row                                                      # Установка индекса 1 листа

    sheet_1.cell(row=user_index_1, column=LAST_DATE_COLUMN).value = '+'                     # Отметка о посещении
    if sheet_1.cell(row=user_index_1, column=USERS_COLUMN + 1).value is None:
        sheet_1.cell(row=user_index_1, column=USERS_COLUMN + 1).value = fio                 # Запись ФИО в 1 лист
        sheet_2.cell(row=user_index_1, column=USERS_COLUMN + 1).value = fio                 # Запись ФИО в 2 лист
        sheet_3.cell(row=user_index_1, column=USERS_COLUMN + 1).value = fio                 # Запись ФИО в 3 лист
        sheet_4.cell(row=user_index_1, column=USERS_COLUMN + 1).value = fio                 # Запись ФИО в 4 лист

    close_table(workbook)


def write_msg(sender, message: str) -> None:
    """
    Функция отправляет сообщение в указанный чат
    :param sender: id чата для отправки сообщения
    :param message: текст сообщения
    :return:
    """
    authorize.method('messages.send', {'chat_id': sender,
                                       'message': message,
                                       'random_id': get_random_id()})


def kick_user(chat_id: int, user_id: int) -> None:
    """
    Функция исключает пользователя по id из указанного чата
    :param chat_id: id чата
    :param user_id: id пользователя
    :return:
    """
    authorize.method('messages.removeChatUser', {"chat_id": chat_id,
                                                 "user_id": user_id,
                                                 "random_id": get_random_id()})


def recording_by_bot() -> None:
    """
    Сценарий работы бота для записи посещаемости на тренировке
    :return:
    """
    print('Бот запущен!\n'
          'Введите в беседу сообщение согласно шаблону: "Записываемся на 01.01.01".\n'
          'При необходимости: "Записываемся на 01.01.01 (тр/сорев/откр)"\n'
          'По окончании записи введите согласно шаблону: "Конец записи".')

    workbook = open_table(False)                                                         # Открытие таблицы
    date = '...'                                                                    # Аварийное значение date

    for event in longpool.listen():                                                 # Инициализация записи на дату
        if event.type == VkBotEventType.MESSAGE_NEW and event.from_chat:
            if event.message.get('from_id') in ADMIN_ID and \
                    event.message.get('text').lower()[:16] == 'записываемся на ':
                date = event.message.get('text').lower()[16:]
                update_table_date(workbook, date)
                write_msg(event.chat_id, 'Начата запись в таблицу!\n'
                                         'Для записи на указанную дату введите "+++ <ФИО> <№ группы>", например, '
                                         '"+++ Иванов Иван Иванович 1234567/12345".\n'
                                         'Если Вы НЕ студент политеха - введите "+++".\n\n'
                                         'ВАЖНО! Пока не появилось сообщение об успешной записи предыдущего человека, '
                                         'свое сообщение для записи не писать!')
                break
    close_table(workbook)

    for event in longpool.listen():                                                 # Процесс записи в таблицу

        if event.type == VkBotEventType.MESSAGE_NEW and event.from_chat:

            if len(event.message.get('text')) >= 3 and event.message.get('text')[:3] == '+++':
                user_get = session_api.users.get(user_ids=event.message.get('from_id'))[0]
                user_name = user_get.get('first_name') + ' ' + user_get.get('last_name')
                if len(event.message.get('text')) > 4:
                    fio = event.message.get('text')[4:]
                else:
                    fio = ''

                update_table_user(event.message.get('from_id'), user_name, fio)
                print(f'{user_name} был на тренировке {date}')
                write_msg(event.chat_id, f'Успешно записано: {user_name} - {date}.')

            if event.message.get('from_id') in ADMIN_ID and event.message.get('text').lower()[:12] == 'конец записи':
                write_msg(event.chat_id, 'Запись окончена!')
                break

    print('Бот выключен!')


def money() -> None:
    """
    Сценарий записи в журнал для донатов
    :return:
    """
    print('Бот запущен!\n'
          'Введите в беседу сообщение согласно шаблону: "Сбор средств от 01.01.01".\n'
          'При необходимости: "Сбор средств от 01.01.01 (расходный материал)"\n'
          'По окончании записи введите согласно шаблону: "Конец записи".')

    workbook = open_table(False)  # Открытие таблицы
    date = '...'  # Аварийное значение date

    for event in longpool.listen():  # Инициализация сбора средств
        if event.type == VkBotEventType.MESSAGE_NEW and event.from_chat:
            if event.message.get('from_id') in ADMIN_ID and \
                    event.message.get('text').lower()[:15] == 'сбор средств от':
                date = 'Сбор средств\n' + event.message.get('text').lower()[16:]
                update_table_date(workbook, date)
                write_msg(event.chat_id, 'Начата запись в таблицу!\n'
                                         'Для подтверждения перевода средств и записи в таблицу '
                                         'пришлите скриншот перевода с сообщением "+++".\n\n'
                                         'ВАЖНО! Пока не появилось сообщение об успешной записи предыдущего человека, '
                                         'свое сообщение для записи не писать!')
                break
    close_table(workbook)

    for event in longpool.listen():  # Процесс записи в таблицу

        if event.type == VkBotEventType.MESSAGE_NEW and event.from_chat:

            if (len(event.message.get('text')) >= 3 and event.message.get('text')[:3] == '+++' and
                len(event.message.get('attachments')) > 0 and
                    event.message.get('attachments')[0].get('type') == 'photo'):

                user_get = session_api.users.get(user_ids=event.message.get('from_id'))[0]
                user_name = user_get.get('first_name') + ' ' + user_get.get('last_name')

                # with open('G:/Мой диск/logs.txt', 'a') as file:                       # Сценарий отладки
                #     file.write(f'\n\nСообщение от {user_name}: {event.message}.')

                update_table_user(event.message.get('from_id'), user_name, '')
                print(f'{user_name} внес средства на {date}')
                write_msg(event.chat_id, f'Успешно записано! {user_name}, благодарим за помощь клубу! ✨')

            if event.message.get('from_id') in ADMIN_ID and event.message.get('text').lower()[:12] == 'конец записи':
                write_msg(event.chat_id, 'Уважаемые участники клуба, благодарим за содействие! ✨ '
                                         'Ресурсы будут направлены на улучшение и '
                                         'поддержание страйкбольного оборудования.')
                break

    print('Бот выключен!')


def kick_illegal_immigrants():
    """
    Сценарий бота для исключения пользователей, не посетивших ни одной тренировки
    :return:
    """
    # Формируем словарь со списком участников чата в формате пары ключ-значение: ID VK - ник юзера
    members = session_api.messages.getConversationMembers(peer_id=2000000000+CHAT_ID_KICK)
    member_ids = {member['id']: f'{member["first_name"]} {member["last_name"]}' for member in members['profiles']}

    workbook = open_table(False)
    sheet_1 = workbook['Посещаемость']
    sheet_5 = workbook['Резерв']
    user_id_list_1 = list(sheet_1.iter_cols(min_col=USERS_COLUMN - 1, max_col=USERS_COLUMN - 1, values_only=True))[0]
    user_id_list_5 = list(sheet_5.iter_cols(min_col=USERS_COLUMN - 1, max_col=USERS_COLUMN - 1, values_only=True))[0]
    legal_ids = user_id_list_1 + user_id_list_5
    close_table(workbook)

    # Соотносим список журнала и список участников беседы
    kick_list = {}
    for person_id, person_name in member_ids.items():
        if person_id not in legal_ids:
            kick_list[person_id] = person_name

    # with open('D:/Мой диск/member_ids.txt', 'w') as file:
    #     for person_id, person_name in kick_list.items():
    #         file.write(f'{person_id} - {person_name}\n')

    write_msg(CHAT_ID_KICK, 'Добрый день!\n'
                            'Участники клуба, не посетившие ни одной тренировки '
                            '(полностью отсутствуют в журнале посещаемости и резерва), '
                            'будут кикнуты из информационной беседы до следующего набора!\n'
                            'Будем ждать Вас в новом сезоне!')

    kick_name_str = ';\n'.join(kick_list.values())
    write_msg(CHAT_ID_KICK, 'Авто-кик:\n'
                            f'{kick_name_str}.')
    print('Авто-кик:\n'
          f'{kick_name_str}.')
    write_msg(CHAT_ID_KICK, 'Для подтверждения отправьте "+++" от имени администратора.')
    for event in longpool.listen():  # Подтверждение списка кика
        if event.type == VkBotEventType.MESSAGE_NEW and event.from_chat:
            if event.message.get('from_id') in ADMIN_ID and \
                    event.message.get('text') == '+++':
                for user_kick in kick_list.keys():
                    kick_user(CHAT_ID_KICK, user_kick)


def day_of_kick():
    """
    Сценарий бота для исключения пользователей по посещаемости последних 5 тренировок
    :return:
    """
    workbook = open_table(False)
    workbook_perc = open_table(True)
    sheet_perc_1 = workbook_perc['Посещаемость']
    sheet_perc_2 = workbook_perc['Соревнования и тренировки']

    sheet_1 = workbook['Посещаемость']
    sheet_2 = workbook['Соревнования и тренировки']
    sheet_3 = workbook['Вооружение и транспорт']
    sheet_4 = workbook['Данные об экипировке']
    sheet_5 = workbook['Резерв']
    sheet_6 = workbook['Кик']

    # Отбираем данные листа с Посещаемостью: Процент посещений, ID VK, Звание, Имя
    percents_list_1 = list(sheet_perc_1.iter_cols(min_col=USERS_COLUMN + 9, max_col=USERS_COLUMN + 9, min_row=2,
                                                  max_row=sheet_perc_1.max_row, values_only=True))[0]
    # TODO: учесть None
    if None in percents_list_1:
        percents_list_1 = percents_list_1[:percents_list_1.index(None)]
    percents_list_1 = list(map(float, percents_list_1))
    user_id_list_1 = list(sheet_1.iter_cols(min_col=USERS_COLUMN - 1, max_col=USERS_COLUMN - 1, min_row=2,
                                            max_row=sheet_1.max_row, values_only=True))[0]
    name_list = list(sheet_1.iter_cols(min_col=USERS_COLUMN, max_col=USERS_COLUMN, min_row=2,
                                       max_row=sheet_1.max_row, values_only=True))[0]
    rank_list = list(sheet_1.iter_cols(min_col=USERS_COLUMN + 2, max_col=USERS_COLUMN + 2, min_row=2,
                                       max_row=sheet_1.max_row, values_only=True))[0]
    percents_dict_1 = {user_id: perc_1 for user_id, perc_1 in zip(user_id_list_1, percents_list_1)}
    name_dict = {user_id: name for user_id, name in zip(user_id_list_1, name_list)}
    rank_dict = {user_id: rank for user_id, rank in zip(user_id_list_1, rank_list)}

    # Отбираем данные листа с Соревнованиями: Процент посещений, ID VK
    percents_list_2 = list(sheet_perc_2.iter_cols(min_col=USERS_COLUMN + 9, max_col=USERS_COLUMN + 9, min_row=2,
                                                  max_row=sheet_perc_2.max_row, values_only=True))[0]
    # TODO: учесть None
    if None in percents_list_2:
        percents_list_2 = percents_list_2[:percents_list_2.index(None)]
    percents_list_2 = list(map(float, percents_list_2))
    user_id_list_2 = list(sheet_2.iter_cols(min_col=USERS_COLUMN - 1, max_col=USERS_COLUMN - 1, min_row=2,
                                            max_row=sheet_2.max_row, values_only=True))[0]
    percents_dict_2 = {user_id: perc_2 for user_id, perc_2 in zip(user_id_list_2, percents_list_2)}

    # Формируем единый словарь с данными: ID VK - [Имя, Звание, Процент 1, Процент 2]
    user_data_dict = {}
    # TODO: учесть None
    for user_id, perc_1 in percents_dict_1.items():
        user_data_dict[user_id] = [name_dict.get(user_id, f"id_{user_id}"), rank_dict.get(user_id, "7.Рекрут"), perc_1, percents_dict_2.get(user_id, 0.0)]

    row_temp = 2                                # Счетчик строки
    kick_id_list, kick_name_list, kick_manual_list, reserve_id_list, reserve_name_list = [], [], [], [], []
    percents_list = list(sheet_perc_1.iter_cols(min_col=USERS_COLUMN + 9, max_col=USERS_COLUMN + 9, min_row=2,
                                                max_row=sheet_perc_1.max_row, values_only=True))[0]
    # TODO: учесть None
    if None in percents_list:
        percents_list = percents_list[:percents_list.index(None)]
    percents_list = list(map(float, percents_list))

    for percent in percents_list:
        if (percent < (PERCENT / 100) and
                user_data_dict[sheet_1.cell(row=row_temp, column=USERS_COLUMN - 1).value][3] < (PERCENT / 100) and
                sheet_1.cell(row=row_temp, column=USERS_COLUMN + 2).value == '7.Рекрут'):                    # Проверка статуса и процента посещаемости за N тренировок и за соревнования

            kick_id_temp = sheet_1.cell(row=row_temp, column=USERS_COLUMN - 1).value
            user_id_list_4 = list(sheet_4.iter_cols(min_col=USERS_COLUMN - 1, max_col=USERS_COLUMN - 1, values_only=True))[0]
            # TODO: учесть None
            if kick_id_temp in user_id_list_4 and sheet_4.cell(row=user_id_list_4.index(kick_id_temp) + 1, column=USERS_COLUMN + 9).value is not None:      # Условие наличия привода
                reserve_id_list.append(int(sheet_1.cell(row=row_temp, column=USERS_COLUMN - 1).value))
                reserve_name_list.append(sheet_1.cell(row=row_temp, column=USERS_COLUMN).value)

                sheet_5.cell(row=sheet_5.max_row + 1, column=USERS_COLUMN - 1).value = \
                    sheet_1.cell(row=row_temp, column=USERS_COLUMN - 1).value  # Перенос ID
                sheet_5.cell(row=sheet_5.max_row, column=USERS_COLUMN).value = \
                    sheet_1.cell(row=row_temp, column=USERS_COLUMN).value  # Перенос юзера
                sheet_5.cell(row=sheet_5.max_row, column=USERS_COLUMN + 1).value = str(datetime.now().date())
            else:
                kick_id_list.append(int(sheet_1.cell(row=row_temp, column=USERS_COLUMN - 1).value))  # Запись в лист ID
                kick_name_list.append(sheet_1.cell(row=row_temp, column=USERS_COLUMN).value)        # Запись имени

                sheet_6.cell(row=sheet_6.max_row + 1, column=USERS_COLUMN - 1).value = \
                    sheet_1.cell(row=row_temp, column=USERS_COLUMN - 1).value  # Перенос ID
                sheet_6.cell(row=sheet_6.max_row, column=USERS_COLUMN).value = \
                    sheet_1.cell(row=row_temp, column=USERS_COLUMN).value  # Перенос юзера
                sheet_6.cell(row=sheet_6.max_row, column=USERS_COLUMN + 1).value = str(datetime.now().date())

            sheet_1.delete_rows(row_temp)  # Удаление из листа 1
            user_id_list_2 = list(sheet_2.iter_cols(min_col=USERS_COLUMN - 1, max_col=USERS_COLUMN - 1, values_only=True))[0]
            if kick_id_temp in user_id_list_2:
                sheet_2.delete_rows(user_id_list_2.index(kick_id_temp) + 1)  # Удаление из листа 2
            user_id_list_3 = list(sheet_3.iter_cols(min_col=USERS_COLUMN - 1, max_col=USERS_COLUMN - 1, values_only=True))[0]
            if kick_id_temp in user_id_list_3:
                sheet_3.delete_rows(user_id_list_3.index(kick_id_temp) + 1)  # Удаление из листа 3
            user_id_list_4 = list(sheet_4.iter_cols(min_col=USERS_COLUMN - 1, max_col=USERS_COLUMN - 1, values_only=True))[0]
            if kick_id_temp in user_id_list_4:
                sheet_4.delete_rows(user_id_list_4.index(kick_id_temp) + 1)  # Удаление из листа 4

            row_temp -= 1  # Шаг строки назад

        row_temp += 1                                                                           # Шаг строки

    close_table(workbook)

    write_msg(CHAT_ID_KICK, 'Добрый день!\n'
                            'Сегодня подведем итоги посещаемости за последние 5 тренировок.\n'
                            f'Участники клуба, посетившие менее {PERCENT}% тренировок, '
                            'будут кикнуты до следующего набора!\n'
                            'Будем ждать Вас в новом сезоне!')

    kick_name_str = ';\n'.join(kick_name_list)
    kick_manual_str = ';\n'.join(kick_manual_list)
    reserve_name_str = ';\n'.join(reserve_name_list)
    write_msg(CHAT_ID_KICK, 'Авто-кик:\n'
                            f'{kick_name_str}\n\n'
                            'Ручной кик:\n'
                            f'{kick_manual_str}.\n\n'
                            'Резерв:\n'
                            f'{reserve_name_str}.\n\n')
    print('Авто-кик:\n'
          f'{kick_name_str}.\n\n'
          'Ручной кик:\n'
          f'{kick_manual_str}.\n\n'
          'Резерв:\n'
          f'{reserve_name_str}.\n\n')
    write_msg(CHAT_ID_KICK, 'Для подтверждения отправьте "+++" от имени администратора.')
    for event in longpool.listen():                                                 # Подтверждение списка кика
        if event.type == VkBotEventType.MESSAGE_NEW and event.from_chat:
            if event.message.get('from_id') in ADMIN_ID and \
                    event.message.get('text') == '+++':
                for user_kick in kick_id_list:
                    kick_user(CHAT_ID_KICK, user_kick)


def listening():
    """
    Сценарий бота для отладки. Прослушивает события в беседе
    :return:
    """
    for event in longpool.listen():
        if event.type == VkBotEventType.MESSAGE_NEW and event.from_chat:
            if event.message.get('from_id') in ADMIN_ID:
                print(event.message)
                print(f'ID: {event.chat_id}')


"""---------------------------------------------Основная программа---------------------------------------------"""


if __name__ == "__main__":
    print('Программа запущена!')
    while True:
        print('Введите номер команды:\n'
              '1 - Запись посещаемости после тренировки.\n'
              '2 - Сбор шекелей.\n'
              '3 - Автокик нелегитимных участников инфо-беседы.\n'
              '4 - День киков!\n'
              '5 - Выход из программы.\n')
        start = input('>>> ')
        try:
            start = int(start)
        except:
            print('Неверная команда, повторите попытку!')
            continue
        match start:
            case 1:
                recording_by_bot()
                print('Команда выполнена!')
            case 2:
                money()
                print('Команда выполнена!')
            case 3:
                kick_illegal_immigrants()
                print('Команда выполнена!')
            case 4:
                day_of_kick()
                print('Команда выполнена!')
            case 5:
                print('Выход из программы!')
                break
            case 999:
                listening()
            case _:
                print('Неверная команда, повторите попытку!')
